#!/usr/bin/env python3
"""Generate a Feishu-style FB weekly report acceptance workbook.

This script intentionally optimizes for the final report blocks shown in the
Feishu weekly report, not for engineer-facing normalized detail sheets.
"""

from __future__ import annotations

from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


ROOT = Path(__file__).resolve().parents[1]
RUN_ROOT = ROOT / "runtime" / "private" / "fb_weekly_v0"
BASE = RUN_ROOT / "fb_weekly_v0_base_table.xlsx"
OUTPUT = RUN_ROOT / "FB自动周报V0_飞书样式验收表.xlsx"
REPORT = ROOT / "docs" / "fb-weekly-v0-feishu-style-gap-report.md"


FEISHU_BLOCKS = {
    "01_FB整体达成": "FB-REDACTED_PRODUCT_LINE -> 整体达成（飞书 sheet: 1AsIod / SkZ0x7）",
    "02_REDACTED_SEGMENT_STRATEGY_DATA": "FB-REDACTED_PRODUCT_LINE -> REDACTED_SEGMENT_STRATEGY_DATA（飞书 sheet: KHAXJJ）",
    "03_REDACTED_CREATIVE_VIEW": "FB-REDACTED_PRODUCT_LINE -> REDACTED_CREATIVE_VIEW / REDACTED_CREATIVE_WASTE_VIEW（飞书 sheet: iuZlU8）",
    "04_REDACTED_RECENT_CREATIVE_VIEW": "FB-REDACTED_PRODUCT_LINE -> REDACTED_RECENT_CREATIVE_VIEW（飞书 sheet: soaSkB）",
    "05_空耗率": "FB-REDACTED_PRODUCT_LINE -> 空耗率（飞书 sheet: tuWkcq）",
    "06_REDACTED_REGION_GROUP数据": "FB-REDACTED_PRODUCT_LINE -> REDACTED_REGION_GROUP 数据（飞书 sheet: VpReZd / Y266Lr / 1Virzt）",
    "07_定位问题与缺口": "本次重做定位：为何上一版没有匹配飞书截图预期",
}


BLUE = "2F5597"
LIGHT_BLUE = "D9EAF7"
ORANGE = "C55A11"
LIGHT_ORANGE = "F8CBAD"
GREEN = "E2F0D9"
YELLOW = "FFF2CC"
GRAY = "E7E6E6"
WHITE = "FFFFFF"
BLACK = "000000"


def clean_column(name: Any) -> str:
    value = str(name)
    value = value.replace(" | nan", "")
    value = value.replace("nan | ", "")
    if " | " in value:
        left, right = value.split(" | ", 1)
        if left == "REDACTED_DIM_OWNER" and right == "消耗":
            return "消耗"
        return right if right and right != "nan" else left
    return "" if value == "nan" else value


def read_sheet(sheet: str) -> pd.DataFrame:
    return pd.read_excel(BASE, sheet_name=sheet, dtype=object)


def to_num(value: Any) -> float | None:
    if value is None or (isinstance(value, float) and pd.isna(value)):
        return None
    if isinstance(value, str):
        value = value.strip().replace(",", "")
        if value.endswith("%"):
            try:
                return float(value[:-1]) / 100
            except ValueError:
                return None
    try:
        number = pd.to_numeric(value, errors="coerce")
    except Exception:
        return None
    return None if pd.isna(number) else float(number)


def fmt_num(value: Any, digits: int = 0) -> Any:
    number = to_num(value)
    if number is None:
        return "" if value is None or (isinstance(value, float) and pd.isna(value)) else value
    if digits == 0:
        return int(round(number))
    return round(number, digits)


def pct(numerator: Any, denominator: Any) -> str:
    n = to_num(numerator)
    d = to_num(denominator)
    if n is None or d in (None, 0):
        return ""
    return f"{n / d:.2%}"


def long_to_wide(df: pd.DataFrame) -> pd.DataFrame:
    source = df.copy()
    source["值"] = source["值"].map(lambda x: to_num(x) if to_num(x) is not None else x)
    wide = (
        source.pivot_table(index=["日期", "分组"], columns="指标", values="值", aggfunc="first")
        .reset_index()
        .rename_axis(None, axis=1)
    )
    return wide


def clean_base(df: pd.DataFrame) -> pd.DataFrame:
    copied = df.copy()
    copied.columns = [clean_column(c) for c in copied.columns]
    copied = copied.loc[:, ~copied.columns.duplicated()]
    hierarchy_cols = [
        "REDACTED_DIM_OWNER",
        "REDACTED_DIM_PLATFORM",
        "REDACTED_DIM_REGION_TIER",
        "投放账户",
        "REDACTED_DIM_FORM_CHANNEL",
        "REDACTED_DIM_TEST_ITEM",
        "REDACTED_FLAG_B",
        "REDACTED_FLAG_A",
    ]
    for col in hierarchy_cols:
        if col in copied.columns:
            copied[col] = copied[col].replace({"": pd.NA, "nan": pd.NA}).ffill()
    return copied


def first_existing(row: pd.Series, names: list[str]) -> Any:
    for name in names:
        if name in row.index and pd.notna(row.get(name)):
            return row.get(name)
    return ""


def safe_col(df: pd.DataFrame, name: str) -> pd.Series:
    if name in df.columns:
        return df[name]
    return pd.Series([None] * len(df), index=df.index)


def aggregate_material(material: pd.DataFrame, group_cols: list[str]) -> pd.DataFrame:
    df = material.copy()
    for col in ["消耗", "REDACTED_CONVERSION_A_COUNT", "REDACTED_CONVERSION_B_COUNT", "当月GMV", "曝光", "点击"]:
        if col in df.columns:
            df[col] = pd.to_numeric(df[col], errors="coerce")
    available = [c for c in group_cols if c in df.columns]
    if not available:
        return pd.DataFrame()
    metrics = [c for c in ["消耗", "REDACTED_CONVERSION_A_COUNT", "REDACTED_CONVERSION_B_COUNT", "当月GMV", "曝光", "点击"] if c in df.columns]
    grouped = df.groupby(available, dropna=False, as_index=False)[metrics].sum()
    if "REDACTED_CONVERSION_A_COUNT" in grouped and "消耗" in grouped:
        grouped["REDACTED_CONVERSION_A_COST"] = grouped.apply(lambda r: r["消耗"] / r["REDACTED_CONVERSION_A_COUNT"] if r["REDACTED_CONVERSION_A_COUNT"] else None, axis=1)
    if "REDACTED_CONVERSION_B_COUNT" in grouped and "消耗" in grouped:
        grouped["REDACTED_CONVERSION_B_COST"] = grouped.apply(lambda r: r["消耗"] / r["REDACTED_CONVERSION_B_COUNT"] if r["REDACTED_CONVERSION_B_COUNT"] else None, axis=1)
    if "REDACTED_CONVERSION_B_COUNT" in grouped and "REDACTED_CONVERSION_A_COUNT" in grouped:
        grouped["REDACTED_CONVERSION_A_TO_B_RATE"] = grouped.apply(lambda r: r["REDACTED_CONVERSION_B_COUNT"] / r["REDACTED_CONVERSION_A_COUNT"] if r["REDACTED_CONVERSION_A_COUNT"] else None, axis=1)
    if "当月GMV" in grouped and "消耗" in grouped:
        grouped["当月ROI2"] = grouped.apply(lambda r: r["当月GMV"] / r["消耗"] if r["消耗"] else None, axis=1)
    if "曝光" in grouped and "消耗" in grouped:
        grouped["CPM"] = grouped.apply(lambda r: r["消耗"] / r["曝光"] * 1000 if r["曝光"] else None, axis=1)
    if "点击" in grouped and "曝光" in grouped:
        grouped["CTR"] = grouped.apply(lambda r: r["点击"] / r["曝光"] if r["曝光"] else None, axis=1)
    return grouped


def channel_summary(channel_wide: pd.DataFrame) -> pd.DataFrame:
    total = channel_wide[channel_wide["日期"].astype(str).eq("总计")].copy()
    if total.empty:
        total = channel_wide.copy()
    for col in ["消耗", "REDACTED_CONVERSION_A_COUNT", "REDACTED_CONVERSION_B_COUNT", "REDACTED_CONVERSION_A_COST", "REDACTED_CONVERSION_B_COST", "REDACTED_CONVERSION_A_TO_B_RATE", "空耗消耗", "空耗占比"]:
        if col in total:
            total[col] = pd.to_numeric(total[col], errors="coerce")
    return total


def build_overall(channel_wide: pd.DataFrame) -> tuple[list[list[Any]], list[int]]:
    total = channel_summary(channel_wide)
    spend = total["消耗"].sum() if "消耗" in total else None
    leads = total["REDACTED_CONVERSION_A_COUNT"].sum() if "REDACTED_CONVERSION_A_COUNT" in total else None
    bookings = total["REDACTED_CONVERSION_B_COUNT"].sum() if "REDACTED_CONVERSION_B_COUNT" in total else None
    lead_cost = spend / leads if spend and leads else None
    booking_cost = spend / bookings if spend and bookings else None
    booking_rate = bookings / leads if bookings and leads else None

    rows: list[list[Any]] = [
        ["对应飞书周报部分", FEISHU_BLOCKS["01_FB整体达成"]],
        ["定位说明", "按截图重建：上方为REDACTED_REGION_AMTD小表，下方为整体达成大表；目标列当前无SmartBI可靠来源，保留待接入。"],
        [],
        ["REDACTED_REGION_A", "MTD目标", "达成", "达成率", "数据状态"],
        ["REDACTED_CONVERSION_A", "目标待接入", fmt_num(leads), "目标待接入", "SmartBI可汇总实际，目标缺失"],
        ["成本", "目标待接入", fmt_num(lead_cost), "目标待接入", "REDACTED_CONVERSION_A_COST=消耗/REDACTED_CONVERSION_A_COUNT，需确认是否含CPT"],
        ["REDACTED_CONVERSION_B_COST", "目标待接入", fmt_num(booking_cost), "目标待接入", "REDACTED_CONVERSION_B_COST=消耗/REDACTED_CONVERSION_B_COUNT，需确认口径"],
        ["REDACTED_DIM_OWNERREDACTED_CONVERSION_B_RATE", "目标待接入", f"{booking_rate:.2%}" if booking_rate is not None else "", "目标待接入", "REDACTED_DIM_OWNER口径待确认"],
        ["REDACTED_DIM_OWNERREDACTED_CONVERSION_B_COUNT", "目标待接入", fmt_num(bookings), "目标待接入", "SmartBI可汇总"],
        ["REDACTED_DIM_OWNER滚动ROI2", "目标待接入", "待接入ROI2", "目标待接入", "渠道日报无ROI2"],
        [],
        ["时间进度：", "待接入日历进度", "", "消耗情况", "REDACTED_CONVERSION_A_COUNT/REDACTED_CONVERSION_B_COUNT情况", "", "", "", "REDACTED_CONVERSION_A_COST情况", "", "", "REDACTED_DIM_OWNERREDACTED_CONVERSION_B情况", "", ""],
        ["REDACTED_DIM_OWNER", "渠道", "渠道/区域", "实际消耗", "REDACTED_DIM_OWNERREDACTED_CONVERSION_AMTD目标", "REDACTED_DIM_OWNERREDACTED_CONVERSION_A_COUNT", "REDACTED_DIM_OWNERREDACTED_CONVERSION_AMTD达成率", "REDACTED_DIM_OWNERREDACTED_CONVERSION_B_COUNT", "REDACTED_DIM_OWNERREDACTED_CONVERSION_A_COST目标", "REDACTED_DIM_OWNERREDACTED_CONVERSION_A_COST", "REDACTED_DIM_OWNERREDACTED_CONVERSION_A_COST目标达成率", "REDACTED_DIM_OWNERREDACTED_CONVERSION_B_RATE目标", "REDACTED_DIM_OWNERREDACTED_CONVERSION_B_RATE", "备注占位"],
    ]
    style_rows = [4, 12, 13]

    region_map = {
        "FBREDACTED_REGION_AKOL": ("REDACTED_REGION_A", "REDACTED_REGION_AKOL"),
        "FBREDACTED_REGION_A常规": ("REDACTED_REGION_A", "REDACTED_REGION_A常规"),
        "REDACTED_PLATFORM_GROUP_A": ("非REDACTED_REGION_A", "REDACTED_REGION_GROUP_A常规"),
        "REDACTED_PLATFORM_GROUP_B": ("非REDACTED_REGION_A", "亚洲常规"),
        "REDACTED_PLATFORM_GROUP_C": ("非REDACTED_REGION_A", "REDACTED_REGION_BKOL"),
        "FBREDACTED_REGION_CKOL": ("非REDACTED_REGION_A", "REDACTED_REGION_CKOL"),
    }
    for _, row in total.iterrows():
        group = str(row.get("分组", ""))
        region, label = region_map.get(group, ("待归类", group))
        rows.append(
            [
                "REDACTED_PRODUCT_LINE",
                "FB",
                region if label == region else label,
                fmt_num(row.get("消耗")),
                "目标待接入",
                fmt_num(row.get("REDACTED_CONVERSION_A_COUNT")),
                "目标待接入",
                fmt_num(row.get("REDACTED_CONVERSION_B_COUNT")),
                "目标待接入",
                fmt_num(row.get("REDACTED_CONVERSION_A_COST")),
                "目标待接入",
                "目标待接入",
                row.get("REDACTED_CONVERSION_A_TO_B_RATE", ""),
                "",
            ]
        )
    rows.append(
        [
            "REDACTED_PRODUCT_LINE",
            "FB",
            "FB汇总",
            fmt_num(spend),
            "目标待接入",
            fmt_num(leads),
            "目标待接入",
            fmt_num(bookings),
            "目标待接入",
            fmt_num(lead_cost),
            "目标待接入",
            "目标待接入",
            f"{booking_rate:.2%}" if booking_rate is not None else "",
            "",
        ]
    )
    return rows, style_rows


def build_hkmo_strategy(hkmo: pd.DataFrame) -> tuple[list[list[Any]], list[int]]:
    df = hkmo.copy()
    columns = ["REDACTED_DIM_REGION_TIER", "REDACTED_DIM_FORM_CHANNEL", "REDACTED_FLAG_B", "REDACTED_FLAG_A", "REDACTED_DIM_TEST_ITEM", "CPM", "CTR", "CVR", "消耗", "REDACTED_CONVERSION_A_COUNT", "REDACTED_CONVERSION_A_COST", "REDACTED_CONVERSION_B_COST", "正价课成本", "REDACTED_CONVERSION_A_TO_B_RATE", "REDACTED_CONVERSION_B_TO_C_RATE", "REDACTED_CONVERSION_C_TO_PAID_RATE", "当月REDACTED_CONVERSION率"]
    rows = [
        ["对应飞书周报部分", FEISHU_BLOCKS["02_REDACTED_SEGMENT_STRATEGY_DATA"]],
        ["定位说明", "对齐截图中的“REDACTED_SEGMENT_STRATEGY_DATA”，保留链路、REDACTED_FLAG_B/REDACTED_FLAG_A、REDACTED_DIM_TEST_ITEM和成本效率字段。"],
        [],
        ["地区", "链路", "REDACTED_FLAG_B", "REDACTED_FLAG_A", "REDACTED_DIM_TEST_ITEM", "cpm", "ctr", "cvr", "消耗", "消耗占比", "REDACTED_CONVERSION_A_COUNT", "REDACTED_CONVERSION_A_COST", "REDACTED_CONVERSION_B_COST", "正价课成本", "REDACTED_CONVERSION_A_TO_B_RATE", "REDACTED_CONVERSION_B_TO_C_RATE", "REDACTED_CONVERSION_C_TO_PAID_RATE", "REDACTED_CONVERSION_A_TO_PAID_RATE", "备注占位"],
    ]
    total_spend = pd.to_numeric(df.get("消耗", pd.Series(dtype=float)), errors="coerce").sum()
    for _, row in df.iterrows():
        if "REDACTED_REGION_A" not in str(first_existing(row, ["REDACTED_DIM_REGION_TIER", "REDACTED_DIM_OWNER_GROUP"])):
            continue
        spend = row.get("消耗")
        spend_display: Any = fmt_num(spend)
        spend_share: Any = pct(spend, total_spend)
        if to_num(spend) in (None, 0) and to_num(row.get("REDACTED_CONVERSION_A_COUNT")) not in (None, 0):
            spend_display = "待校准"
            spend_share = "待校准"
        rows.append(
            [
                "REDACTED_REGION_A",
                row.get("REDACTED_DIM_FORM_CHANNEL", ""),
                row.get("REDACTED_FLAG_B", ""),
                row.get("REDACTED_FLAG_A", ""),
                row.get("REDACTED_DIM_TEST_ITEM", ""),
                row.get("CPM", ""),
                row.get("CTR", ""),
                row.get("CVR", ""),
                spend_display,
                spend_share,
                fmt_num(row.get("REDACTED_CONVERSION_A_COUNT")),
                fmt_num(row.get("REDACTED_CONVERSION_A_COST")),
                fmt_num(row.get("REDACTED_CONVERSION_B_COST")),
                fmt_num(row.get("正价课成本")),
                row.get("REDACTED_CONVERSION_A_TO_B_RATE", ""),
                row.get("REDACTED_CONVERSION_B_TO_C_RATE", ""),
                row.get("REDACTED_CONVERSION_C_TO_PAID_RATE", ""),
                row.get("当月REDACTED_CONVERSION率", ""),
                "",
            ]
        )
    return rows, [4]


def build_material(material: pd.DataFrame) -> tuple[list[list[Any]], list[int]]:
    df = material.copy()
    for col in ["消耗", "REDACTED_CONVERSION_A_COUNT", "REDACTED_CONVERSION_A_COST", "REDACTED_CONVERSION_B_COST"]:
        if col in df:
            df[col] = pd.to_numeric(df[col], errors="coerce")
    df = df[df.get("消耗", pd.Series(0, index=df.index)).fillna(0).gt(0)].copy()
    df = df[df.get("广告名称", pd.Series("", index=df.index)).astype(str).ne("总计")]
    df = df.sort_values("消耗", ascending=False).head(60)
    rows = [
        ["对应飞书周报部分", FEISHU_BLOCKS["03_REDACTED_CREATIVE_VIEW"]],
        ["定位说明", "对齐截图中的“REDACTED_CREATIVE_VIEW/REDACTED_CREATIVE_WASTE_VIEW”，不是输出所有creative_asset明细；按消耗取Topcreative_asset供周报展示。"],
        [],
        ["地区", "REDACTED_CREATIVE_TYPE", "广告名称", "cpm", "ctr", "cvr", "消耗", "REDACTED_CONVERSION_A_COUNT", "REDACTED_CONVERSION_A_COST", "REDACTED_CONVERSION_B_COST", "REDACTED_CONVERSION_A_TO_B_RATE", "REDACTED_CONVERSION_B_TO_C_RATE", "REDACTED_CONVERSION_C_TO_PAID_RATE", "REDACTED_CONVERSION_A_TO_PAID_RATE", "当月ROI2", "备注占位"],
    ]
    for _, row in df.iterrows():
        rows.append(
            [
                row.get("REDACTED_DIM_REGION_TIER", ""),
                row.get("REDACTED_CREATIVE_TYPE", ""),
                row.get("广告名称", ""),
                row.get("CPM", ""),
                row.get("CTR", ""),
                row.get("CVR", ""),
                fmt_num(row.get("消耗")),
                fmt_num(row.get("REDACTED_CONVERSION_A_COUNT")),
                fmt_num(row.get("REDACTED_CONVERSION_A_COST")),
                fmt_num(row.get("REDACTED_CONVERSION_B_COST")),
                row.get("REDACTED_CONVERSION_B_RATE", ""),
                row.get("REDACTED_CONVERSION_B_TO_C_RATE", ""),
                row.get("REDACTED_CONVERSION_C_TO_PAID_RATE", ""),
                row.get("当月REDACTED_CONVERSION率", ""),
                row.get("当月ROI2", ""),
                "",
            ]
        )
    return rows, [4]


def build_recent_material(material: pd.DataFrame) -> tuple[list[list[Any]], list[int]]:
    df = material.copy()
    fallback = df.copy()
    for col in ["消耗", "REDACTED_CONVERSION_A_COUNT"]:
        if col in df:
            df[col] = pd.to_numeric(df[col], errors="coerce")
            fallback[col] = pd.to_numeric(fallback[col], errors="coerce")
    if "上线日期" in df:
        df = df[df["上线日期"].notna()].copy()
        df = df[df["上线日期"].astype(str).ne("总计")]
    for col in ["REDACTED_DIM_REGION_TIER", "REDACTED_CREATIVE_TYPE", "广告名称"]:
        if col in df:
            df = df[df[col].astype(str).ne("总计")]
    df = df[df.get("消耗", pd.Series(0, index=df.index)).fillna(0).gt(0)].copy()
    if df.empty:
        df = fallback.copy()
        for col in ["REDACTED_DIM_REGION_TIER", "REDACTED_CREATIVE_TYPE", "广告名称"]:
            if col in df:
                df = df[~df[col].astype(str).isin(["总计", "nan", "None", ""])]
        df = df[df.get("消耗", pd.Series(0, index=df.index)).fillna(0).gt(0)].copy()
    if "上线日期" in df:
        df = df.sort_values(["上线日期", "消耗"], ascending=[False, False])
    else:
        df = df.sort_values("消耗", ascending=False)
    df = df.head(50)
    rows = [
        ["对应飞书周报部分", FEISHU_BLOCKS["04_REDACTED_RECENT_CREATIVE_VIEW"]],
        ["定位说明", "对齐截图中的“REDACTED_RECENT_CREATIVE_VIEW”；SmartBIcreative_asset表缺少REDACTED_CREATIVE_PRODUCTION_PERIOD、KOL、预览链接，先保留待接入栏。"],
        [],
        ["REDACTED_CREATIVE_PRODUCTION_PERIOD", "KOL", "上线日期", "预览链接", "投放地区", "REDACTED_CREATIVE_TYPE", "广告名称", "cpm", "ctr", "cvr", "消耗", "REDACTED_CONVERSION_A_COUNT", "REDACTED_CONVERSION_A_COST", "REDACTED_CONVERSION_B_COST", "REDACTED_CONVERSION_A_TO_B_RATE", "REDACTED_CONVERSION_B_TO_C_RATE", "REDACTED_CONVERSION_C_TO_PAID_RATE", "REDACTED_CONVERSION_A_TO_PAID_RATE", "25%播放进度", "50%播放进度", "备注占位"],
    ]
    for _, row in df.iterrows():
        rows.append(
            [
                "待接入creative_asset产出表",
                "待接入KOL",
                row.get("上线日期", ""),
                "待接入预览链接",
                row.get("REDACTED_DIM_REGION_TIER", ""),
                row.get("REDACTED_CREATIVE_TYPE", ""),
                row.get("广告名称", ""),
                row.get("CPM", ""),
                row.get("CTR", ""),
                row.get("CVR", ""),
                fmt_num(row.get("消耗")),
                fmt_num(row.get("REDACTED_CONVERSION_A_COUNT")),
                fmt_num(row.get("REDACTED_CONVERSION_A_COST")),
                fmt_num(row.get("REDACTED_CONVERSION_B_COST")),
                row.get("REDACTED_CONVERSION_B_RATE", ""),
                row.get("REDACTED_CONVERSION_B_TO_C_RATE", ""),
                row.get("REDACTED_CONVERSION_C_TO_PAID_RATE", ""),
                row.get("当月REDACTED_CONVERSION率", ""),
                row.get("完播25%", ""),
                row.get("完播50%", ""),
                "",
            ]
        )
    return rows, [4]


def build_waste(material: pd.DataFrame) -> tuple[list[list[Any]], list[int]]:
    df = material.copy()
    for col in ["消耗", "成效", "成效成本", "空耗金额", "空耗金额占比"]:
        if col in df:
            df[col] = pd.to_numeric(df[col], errors="coerce")
    rows = [
        ["对应飞书周报部分", FEISHU_BLOCKS["05_空耗率"]],
        ["定位说明", "飞书原表是账户/周次/新老计划/广告组/广告维度结构；SmartBIcreative_asset表目前只能支撑广告/creative_asset级候选，缺少新老计划拆分。"],
        [],
        ["账户类型", "周次", "汇总-消耗", "汇总-成效", "汇总-成效成本", "新计划-消耗", "新计划-成效", "新计划-成效成本", "老计划-消耗", "老计划-成效", "老计划-成效成本", "广告组维度-空耗金额占比", "广告维度-空耗金额占比", "候选类型", "人工确认位"],
    ]
    group_cols = [c for c in ["REDACTED_DIM_PLATFORM", "REDACTED_DIM_REGION_TIER"] if c in df.columns]
    if group_cols:
        for col in group_cols:
            df = df[~df[col].astype(str).isin(["总计", "nan", "None", ""])]
        grouped = df.groupby(group_cols, dropna=False).agg({"消耗": "sum", "成效": "sum", "空耗金额": "sum"}).reset_index()
        grouped["成效成本"] = grouped.apply(lambda r: r["消耗"] / r["成效"] if r["成效"] else None, axis=1)
        grouped["空耗金额占比"] = grouped.apply(lambda r: r["空耗金额"] / r["消耗"] if r["消耗"] else None, axis=1)
        grouped = grouped.sort_values(["空耗金额", "消耗"], ascending=False).head(30)
        for _, row in grouped.iterrows():
            label = " / ".join(str(row.get(c, "")) for c in group_cols)
            rows.append(
                [
                    label,
                    "本期SmartBI导出窗口",
                    fmt_num(row.get("消耗")),
                    fmt_num(row.get("成效")),
                    fmt_num(row.get("成效成本")),
                    "待接入新老计划字段",
                    "待接入新老计划字段",
                    "待接入新老计划字段",
                    "待接入新老计划字段",
                    "待接入新老计划字段",
                    "待接入新老计划字段",
                    "待接入广告组空耗口径",
                    row.get("空耗金额占比", ""),
                    "排序候选，不给处理建议",
                    "",
                ]
            )
    return rows, [4]


def build_region_sg(material: pd.DataFrame, channel_wide: pd.DataFrame) -> tuple[list[list[Any]], list[int]]:
    total = channel_summary(channel_wide)
    rows = [
        ["对应飞书周报部分", FEISHU_BLOCKS["06_REDACTED_REGION_GROUP数据"]],
        ["定位说明", "对齐飞书中的REDACTED_REGION_GROUP_A/REDACTED_REGION_B常规/REDACTED_REGION_BKOL目标小表和REDACTED_CREATIVE_TYPE表；目标与部分地域拆分仍待接入。"],
        [],
        ["REDACTED_REGION_GROUP_A", "", "", "", "", "REDACTED_REGION_B常规", "", "", "", "", "REDACTED_REGION_BKOL", "", "", ""],
        ["指标", "目标", "达成", "达成率", "", "指标", "目标", "达成", "达成率", "", "指标", "目标", "达成", "达成率"],
    ]
    blocks = [
        ("REDACTED_PLATFORM_GROUP_A", "REDACTED_REGION_GROUP_A"),
        ("REDACTED_PLATFORM_GROUP_B", "REDACTED_REGION_B常规"),
        ("REDACTED_PLATFORM_GROUP_C", "REDACTED_REGION_BKOL"),
    ]
    metrics = [("REDACTED_CONVERSION_A", "REDACTED_CONVERSION_A_COUNT"), ("成本", "REDACTED_CONVERSION_A_COST"), ("REDACTED_CONVERSION_B_COST", "REDACTED_CONVERSION_B_COST"), ("REDACTED_DIM_OWNERREDACTED_CONVERSION_B_RATE", "REDACTED_CONVERSION_A_TO_B_RATE"), ("REDACTED_DIM_OWNERREDACTED_CONVERSION_B_COUNT", "REDACTED_CONVERSION_B_COUNT"), ("REDACTED_DIM_OWNER滚动ROI2", None)]
    for metric_label, col in metrics:
        row_out: list[Any] = []
        for group, _label in blocks:
            source = total[total["分组"].astype(str).eq(group)]
            value = "" if source.empty or col is None or col not in source.columns else source.iloc[0].get(col)
            row_out.extend([metric_label, "目标待接入", fmt_num(value) if col != "REDACTED_CONVERSION_A_TO_B_RATE" else value, "目标待接入", ""])
        rows.append(row_out[:-1])

    rows.extend(
        [
            [],
            ["REDACTED_REGION_GROUP_AREDACTED_CREATIVE_TYPE数据：", "SmartBIcreative_asset表聚合，需确认是否等于飞书REDACTED_CREATIVE_TYPE口径"],
            ["地区", "REDACTED_CREATIVE_TYPE", "CTR", "CPM", "CVR", "消耗", "REDACTED_CONVERSION_A_COUNT", "REDACTED_CONVERSION_A_COST", "REDACTED_CONVERSION_B_COST", "REDACTED_CONVERSION_A_TO_B_RATE", "REDACTED_CONVERSION_C_TO_PAID_RATE", "当月ROI2"],
        ]
    )
    grouped = aggregate_material(material, ["REDACTED_DIM_REGION_TIER", "REDACTED_CREATIVE_TYPE"])
    if not grouped.empty:
        mask = grouped["REDACTED_DIM_REGION_TIER"].astype(str).str.contains("REDACTED_REGION_GROUP_A|REDACTED_REGION_B|非REDACTED_REGION_A|亚洲", na=False)
        grouped = grouped[mask].sort_values("消耗", ascending=False).head(40)
        for _, row in grouped.iterrows():
            rows.append(
                [
                    row.get("REDACTED_DIM_REGION_TIER", ""),
                    row.get("REDACTED_CREATIVE_TYPE", ""),
                    row.get("CTR", ""),
                    fmt_num(row.get("CPM")),
                    row.get("CVR", ""),
                    fmt_num(row.get("消耗")),
                    fmt_num(row.get("REDACTED_CONVERSION_A_COUNT")),
                    fmt_num(row.get("REDACTED_CONVERSION_A_COST")),
                    fmt_num(row.get("REDACTED_CONVERSION_B_COST")),
                    row.get("REDACTED_CONVERSION_A_TO_B_RATE", ""),
                    "待接入/待确认",
                    row.get("当月ROI2", ""),
                ]
            )
    return rows, [4, 5, 14]


def build_gaps() -> tuple[list[list[Any]], list[int]]:
    rows = [
        ["对应飞书周报部分", FEISHU_BLOCKS["07_定位问题与缺口"]],
        ["定位结论", "上一版没有匹配预期，不是 SmartBI 数据导出失败，而是交付视角错了：输出了工程细表/宽表，没有按飞书最终汇报块重构。"],
        [],
        ["问题", "上一版表现", "根因", "本版处理", "仍需校准"],
        ["整体达成不直观", "Sheet里是区域/策略长表，不像截图的MTD小表+大汇总表", "把 SmartBI 报表当分析底表，没有把飞书 sheet 1AsIod / SkZ0x7 当模板", "新增 01_FB整体达成，按截图结构输出", "目标表、时间进度、ROI2来源"],
        ["REDACTED_REGION_A分策略缺位", "被放在“FBREDACTED_SEGMENT_TEST”，字段也偏底表", "未把 KHAXJJ 识别为独立展示块", "新增 02_REDACTED_SEGMENT_STRATEGY_DATA，按地区/链路/REDACTED_FLAG_B/REDACTED_FLAG_A输出", "消耗占比和REDACTED_DIM_TEST_ITEM口径"],
        ["REDACTED_CREATIVE_VIEW不像周报", "输出 Top100 creative_asset细表，字段过多", "没有按 iuZlU8 的creative_asset表现表裁剪字段", "新增 03_REDACTED_CREATIVE_VIEW，只保留周报展示字段", "REDACTED_CREATIVE_TYPE、广告名称、ROI2口径"],
        ["REDACTED_RECENT_CREATIVE_VIEW缺失", "上一版没有单独展示", "SmartBIcreative_asset表缺少REDACTED_CREATIVE_PRODUCTION_PERIOD/KOL/预览链接，但应显式占位", "新增 04_REDACTED_RECENT_CREATIVE_VIEW", "REDACTED_CREATIVE_ASSET_TABLE或链接表接入"],
        ["空耗率不匹配", "上一版是广告候选列表，不是周次/账户结构", "缺少飞书 tuWkcq 的账户周次结构和新老计划字段", "新增 05_空耗率，按飞书结构标注不可自动填字段", "新老计划、广告组空耗阈值"],
        ["REDACTED_REGION_GROUP遗漏", "只有混在creative_asset表现/REDACTED_REGION_A测试中", "未按 VpReZd/Y266Lr/1Virzt 分块", "新增 06_REDACTED_REGION_GROUP数据", "地区映射、目标、语言维度"],
        [],
        ["字段/数据缺口", "当前状态", "影响", "是否阻塞V0", "建议补齐来源"],
        ["目标字段", "SmartBI四张已导出报表未发现可靠目标表", "达成率无法自动计算", "不阻塞展示骨架，阻塞自动达成率", "目标配置表/投放目标报表"],
        ["时间进度", "未从BI接入自然月/周进度", "整体达成大表顶部无法自动填", "不阻塞", "日期维表或脚本计算"],
        ["ROI2", "creative_asset表有当月/滚动ROI2，渠道日报没有", "整体达成/区域目标表不完整", "不阻塞", "GMV/ROI报表或creative_asset表聚合口径确认"],
        ["预览链接/KOL/REDACTED_CREATIVE_PRODUCTION_PERIOD", "SmartBIcreative_asset维度表缺失", "REDACTED_RECENT_CREATIVE_VIEW无法完整自动生成", "不阻塞", "creative_asset资产管理表"],
        ["新老计划/广告组空耗", "当前creative_asset表只有空耗金额字段", "空耗率表无法复刻", "不阻塞", "空耗专项报表或广告组计划标签"],
    ]
    return rows, [4, 13]


def write_block(writer: pd.ExcelWriter, name: str, rows: list[list[Any]], style_rows: list[int]) -> None:
    df = pd.DataFrame(rows)
    df.to_excel(writer, sheet_name=name, index=False, header=False)
    ws = writer.book[name]
    ws.freeze_panes = "A4"
    ws.sheet_view.showGridLines = False
    thin = Side(style="thin", color="BFBFBF")
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
            if cell.value is None:
                cell.border = Border()
    for row_idx in style_rows:
        if row_idx <= ws.max_row:
            for cell in ws[row_idx]:
                if cell.value is not None:
                    cell.fill = PatternFill("solid", fgColor=BLUE)
                    cell.font = Font(color=WHITE, bold=True)
                    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for cell in ws[1]:
        if cell.value is not None:
            cell.fill = PatternFill("solid", fgColor=LIGHT_BLUE)
            cell.font = Font(bold=True)
    for row in ws.iter_rows():
        values = [str(c.value) for c in row if c.value is not None]
        if any("目标待接入" in v or "待接入" in v or "待校准" in v for v in values):
            for cell in row:
                if cell.value is not None:
                    cell.fill = PatternFill("solid", fgColor=YELLOW)
        if any(str(c.value).endswith("汇总") or str(c.value).endswith("合计") for c in row if c.value is not None):
            for cell in row:
                if cell.value is not None:
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill("solid", fgColor=LIGHT_BLUE)
    for idx, column_cells in enumerate(ws.columns, start=1):
        max_len = 0
        for cell in list(column_cells)[:120]:
            if cell.value is not None:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[get_column_letter(idx)].width = min(max(max_len + 2, 10), 28)
    for r in range(1, ws.max_row + 1):
        ws.row_dimensions[r].height = 24


def write_markdown() -> None:
    lines = [
        "# FB 自动周报 V0 飞书样式验收表说明",
        "",
        "生成日期：2026-05-18",
        "",
        "## 当前产物",
        "",
        f"- Excel：`{OUTPUT.relative_to(ROOT)}`",
        "- 本版目标：直接按飞书 FB 周报最终展示块生成验收表，不再让业务从工程细表里筛选。",
        "",
        "## 对应飞书位置",
        "",
        "| Sheet | 对应飞书周报部分 | 状态 |",
        "|---|---|---|",
    ]
    status = {
        "01_FB整体达成": "结构已对齐；目标、时间进度、ROI2待接入。",
        "02_REDACTED_SEGMENT_STRATEGY_DATA": "结构已对齐；使用 `REDACTED_REPORT_D` 填充REDACTED_REGION_A策略数据。",
        "03_REDACTED_CREATIVE_VIEW": "结构已对齐；使用 `REDACTED_REPORT_A` 按消耗Top展示。",
        "04_REDACTED_RECENT_CREATIVE_VIEW": "结构已对齐；REDACTED_CREATIVE_PRODUCTION_PERIOD/KOL/预览链接待接入REDACTED_CREATIVE_ASSET_TABLE。",
        "05_空耗率": "结构已对齐；新老计划/广告组维度空耗字段待接入。",
        "06_REDACTED_REGION_GROUP数据": "结构已对齐；区域映射、目标、语言维度待校准。",
        "07_定位问题与缺口": "解释上一版未达预期的原因和本版修正点。",
    }
    for sheet, part in FEISHU_BLOCKS.items():
        lines.append(f"| `{sheet}` | {part} | {status[sheet]} |")
    lines.extend(
        [
            "",
            "## 上一版没有匹配预期的原因",
            "",
            "1. 输出视角错了：上一版是 SmartBI 底表整理，不是飞书最终汇报块。",
            "2. 模板锚点错了：没有把飞书正文中的 `1AsIod / SkZ0x7 / KHAXJJ / iuZlU8 / soaSkB / tuWkcq / VpReZd` 作为目标结构。",
            "3. 字段分组错了：整体达成、REDACTED_REGION_A分策略、REDACTED_CREATIVE_VIEW、REDACTED_RECENT_CREATIVE_VIEW、空耗率、REDACTED_REGION_GROUP应拆成独立汇报块。",
            "4. 缺口表达错了：缺少目标、ROI2、creative_asset链接、新老计划等字段时，上一版没有在最终展示结构中显式标注。",
            "",
            "## 需要你验收/校准",
            "",
            "- `目标字段`：MTD目标、达成率、时间进度是否来自另一个目标表。",
            "- `区域映射`：SmartBI分组里的 `REDACTED_PLATFORM_GROUP_A/REDACTED_PLATFORM_GROUP_B/REDACTED_PLATFORM_GROUP_C/FBREDACTED_REGION_CKOL` 是否对应飞书的REDACTED_REGION_GROUP_A、REDACTED_REGION_B常规、REDACTED_REGION_BKOL等块。",
            "- `消耗口径`：周报采用消耗、消耗不含CPT，还是含CPT。",
            "- `ROI2口径`：整体达成使用当月ROI2还是滚动ROI2。",
            "- `REDACTED_CREATIVE_ASSET_FIELDS`：REDACTED_RECENT_CREATIVE_VIEW需要的 KOL、REDACTED_CREATIVE_PRODUCTION_PERIOD、预览链接是否有独立creative_asset表。",
            "- `空耗口径`：新计划/老计划、广告组维度/广告维度空耗率是否有专项BI报表。",
        ]
    )
    REPORT.write_text("\n".join(lines) + "\n", encoding="utf-8")


def main() -> None:
    if not BASE.exists():
        raise SystemExit(f"Base workbook not found: {BASE}")
    channel_wide = long_to_wide(read_sheet("channel_daily_long"))
    material = clean_base(read_sheet("material_wide"))
    hkmo = clean_base(read_sheet("hkmo_strategy_wide"))

    builders = {
        "01_FB整体达成": build_overall(channel_wide),
        "02_REDACTED_SEGMENT_STRATEGY_DATA": build_hkmo_strategy(hkmo),
        "03_REDACTED_CREATIVE_VIEW": build_material(material),
        "04_REDACTED_RECENT_CREATIVE_VIEW": build_recent_material(material),
        "05_空耗率": build_waste(material),
        "06_REDACTED_REGION_GROUP数据": build_region_sg(material, channel_wide),
        "07_定位问题与缺口": build_gaps(),
    }
    OUTPUT.parent.mkdir(parents=True, exist_ok=True)
    with pd.ExcelWriter(OUTPUT, engine="openpyxl") as writer:
        for name, (rows, style_rows) in builders.items():
            write_block(writer, name, rows, style_rows)
    write_markdown()
    print(OUTPUT)
    print(REPORT)


if __name__ == "__main__":
    main()
