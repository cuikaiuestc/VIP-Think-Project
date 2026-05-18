#!/usr/bin/env python3
"""Generate business-facing FB weekly V0 acceptance workbook."""

from __future__ import annotations

from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill


ROOT = Path(__file__).resolve().parents[1]
RUN_ROOT = ROOT / "runtime" / "private" / "fb_weekly_v0"
BASE = RUN_ROOT / "fb_weekly_v0_base_table.xlsx"
OUTPUT = RUN_ROOT / "FB自动周报V0_验收表.xlsx"
REPORT = ROOT / "docs" / "fb-weekly-v0-acceptance-report.md"


FEISHU = {
    "目录_对应飞书位置": "飞书周报目录对照",
    "FB整体达成": "FB-REDACTED_PRODUCT_LINE -> 整体达成",
    "FB渠道日监控": "FB-REDACTED_PRODUCT_LINE -> 整体达成 / REDACTED_SEGMENT_STRATEGY_DATA",
    "FB链路类型日监控": "FB-REDACTED_PRODUCT_LINE -> REDACTED_SEGMENT_STRATEGY_DATA",
    "FBcreative_asset表现": "FB-REDACTED_PRODUCT_LINE -> REDACTED_CREATIVE_VIEW / REDACTED_RECENT_CREATIVE_VIEW / REDACTED_REGION_GROUP 数据 -> REDACTED_CREATIVE_TYPE数据",
    "FB空耗高成本候选": "FB-REDACTED_PRODUCT_LINE -> REDACTED_CREATIVE_VIEW -> REDACTED_CREATIVE_WASTE_VIEW / 空耗率",
    "FBREDACTED_SEGMENT_TEST": "FB-REDACTED_PRODUCT_LINE -> REDACTED_SEGMENT_STRATEGY_DATA / REDACTED_REGION_GROUP 数据",
    "数据缺口与人工补充": "周报人工观点和口径补充区",
}


def clean_column(name: str) -> str:
    value = str(name)
    value = value.replace(" | nan", "")
    value = value.replace("nan | ", "")
    if " | " in value:
        left, right = value.split(" | ", 1)
        if left in {"REDACTED_DIM_OWNER"} and right == "消耗":
            return "消耗"
        if right and right != left:
            return right
        return left
    return "" if value == "nan" else value


def to_number(series: pd.Series) -> pd.Series:
    return pd.to_numeric(series, errors="coerce")


def read_base(sheet: str) -> pd.DataFrame:
    return pd.read_excel(BASE, sheet_name=sheet, dtype=object)


def add_context(df: pd.DataFrame, sheet_name: str) -> pd.DataFrame:
    copied = df.copy()
    copied.insert(0, "对应飞书周报部分", FEISHU[sheet_name])
    return copied


def long_to_wide(df: pd.DataFrame) -> pd.DataFrame:
    source = df.copy()
    numeric_values = pd.to_numeric(source["值"], errors="coerce")
    source["值"] = source["值"].where(numeric_values.isna(), numeric_values)
    wide = (
        source.pivot_table(index=["日期", "分组"], columns="指标", values="值", aggfunc="first")
        .reset_index()
        .rename_axis(None, axis=1)
    )
    return wide


def pick_columns(df: pd.DataFrame, columns: list[str]) -> pd.DataFrame:
    result = pd.DataFrame()
    for column in columns:
        result[column] = df[column] if column in df.columns else ""
    return result


def directory_sheet() -> pd.DataFrame:
    rows = [
        {
            "本Excel表名": "FB整体达成",
            "对应飞书周报部分": FEISHU["FB整体达成"],
            "使用SmartBI报表": "REDACTED_REPORT_B",
            "主要用途": "按分组汇总消耗、REDACTED_CONVERSION_A、REDACTED_CONVERSION_B等核心指标；目标列暂留空。",
            "是否可直接用于V0": "部分可用",
            "待校准点": "目标字段来源、达成率口径、REDACTED_CONVERSION_C/REDACTED_PAID_EVENT/GMV/ROI2是否从其他报表补齐。",
        },
        {
            "本Excel表名": "FB渠道日监控",
            "对应飞书周报部分": FEISHU["FB渠道日监控"],
            "使用SmartBI报表": "REDACTED_REPORT_B",
            "主要用途": "按日期和FB分组展示CPM、CTR、CVR、消耗、REDACTED_CONVERSION_A、REDACTED_CONVERSION_B、空耗。",
            "是否可直接用于V0": "可用",
            "待校准点": "分组命名是否与周报正式分组一致；消耗是否含CPT。",
        },
        {
            "本Excel表名": "FB链路类型日监控",
            "对应飞书周报部分": FEISHU["FB链路类型日监控"],
            "使用SmartBI报表": "REDACTED_REPORT_C",
            "主要用途": "按H5、WhatsApp、表单等链路展示日监控指标。",
            "是否可直接用于V0": "可用",
            "待校准点": "链路类型是否只覆盖REDACTED_REGION_A；是否需要非REDACTED_REGION_A链路。",
        },
        {
            "本Excel表名": "FBcreative_asset表现",
            "对应飞书周报部分": FEISHU["FBcreative_asset表现"],
            "使用SmartBI报表": "REDACTED_REPORT_A",
            "主要用途": "展示Topcreative_asset/广告表现，供周报creative_asset部分验收。",
            "是否可直接用于V0": "可用",
            "待校准点": "REDACTED_CREATIVE_TYPE、REDACTED_DIM_PLATFORM、广告对象字段是否为正式口径。",
        },
        {
            "本Excel表名": "FB空耗高成本候选",
            "对应飞书周报部分": FEISHU["FB空耗高成本候选"],
            "使用SmartBI报表": "REDACTED_REPORT_A",
            "主要用途": "按排序列出空耗金额、高成效成本、高消耗低成效候选。",
            "是否可直接用于V0": "部分可用",
            "待校准点": "空耗阈值、成效口径、候选类型规则。",
        },
        {
            "本Excel表名": "FBREDACTED_SEGMENT_TEST",
            "对应飞书周报部分": FEISHU["FBREDACTED_SEGMENT_TEST"],
            "使用SmartBI报表": "REDACTED_REPORT_D",
            "主要用途": "展示区域、链路、REDACTED_FLAG_B/REDACTED_FLAG_A、creative_asset策略的表现。",
            "是否可直接用于V0": "部分可用",
            "待校准点": "是否纳入V0，或作为V1策略分析模块。",
        },
        {
            "本Excel表名": "数据缺口与人工补充",
            "对应飞书周报部分": FEISHU["数据缺口与人工补充"],
            "使用SmartBI报表": "全部",
            "主要用途": "集中列出口径缺口和人工填写位。",
            "是否可直接用于V0": "可用",
            "待校准点": "确认人和最终口径。",
        },
    ]
    return pd.DataFrame(rows)


def overall_sheet(channel_wide: pd.DataFrame) -> pd.DataFrame:
    total = channel_wide[channel_wide["日期"].astype(str) == "总计"].copy()
    if total.empty:
        total = channel_wide.groupby("分组", as_index=False).agg(
            {"消耗": "sum", "REDACTED_CONVERSION_A_COUNT": "sum", "REDACTED_CONVERSION_B_COUNT": "sum"}
        )
    rows = []
    for _, row in total.iterrows():
        rows.append(
            {
                "区域/策略": row.get("分组", ""),
                "指标": "汇总",
                "目标": "目标待接入",
                "实际": "见右侧核心指标",
                "达成率": "目标待接入",
                "消耗": row.get("消耗", ""),
                "REDACTED_CONVERSION_A_COUNT": row.get("REDACTED_CONVERSION_A_COUNT", ""),
                "REDACTED_CONVERSION_B_COUNT": row.get("REDACTED_CONVERSION_B_COUNT", ""),
                "REDACTED_CONVERSION_C_COUNT": "",
                "REDACTED_PAID_EVENT/REDACTED_PAID_EVENT": "",
                "GMV": "",
                "ROI2": "",
                "备注占位": "",
            }
        )
    return add_context(pd.DataFrame(rows), "FB整体达成")


def channel_sheet(channel_wide: pd.DataFrame) -> pd.DataFrame:
    df = channel_wide[channel_wide["日期"].astype(str) != "总计"].copy()
    columns = ["日期", "分组", "CPM", "CTR", "CVR", "消耗", "REDACTED_CONVERSION_A_COUNT", "REDACTED_CONVERSION_A_COST", "REDACTED_CONVERSION_B_COUNT", "REDACTED_CONVERSION_B_COST", "REDACTED_CONVERSION_A_TO_B_RATE", "空耗消耗", "空耗占比"]
    result = pick_columns(df, columns)
    result["备注占位"] = ""
    return add_context(result, "FB渠道日监控")


def link_type_sheet(link_wide: pd.DataFrame) -> pd.DataFrame:
    df = link_wide[link_wide["日期"].astype(str) != "汇总"].copy()
    df = df.rename(columns={"分组": "链路类型"})
    columns = ["日期", "链路类型", "CPM", "CTR", "CVR", "消耗", "REDACTED_CONVERSION_A_COUNT", "REDACTED_CONVERSION_B_COUNT", "REDACTED_CONVERSION_A_COST", "REDACTED_CONVERSION_B_RATE", "REDACTED_CONVERSION_B_COST"]
    result = pick_columns(df, columns)
    result["备注占位"] = ""
    return add_context(result, "FB链路类型日监控")


def material_sheet(material: pd.DataFrame) -> pd.DataFrame:
    df = material.copy()
    df.columns = [clean_column(c) for c in df.columns]
    df = df.loc[:, ~df.columns.duplicated()]
    for col in ["消耗", "REDACTED_CONVERSION_A_COUNT", "REDACTED_CONVERSION_B_COUNT"]:
        if col in df.columns:
            df[col] = to_number(df[col])
    mask = pd.Series(False, index=df.index)
    for col in ["消耗", "REDACTED_CONVERSION_A_COUNT", "REDACTED_CONVERSION_B_COUNT"]:
        if col in df.columns:
            mask = mask | (df[col].fillna(0) > 0)
    df = df[mask]
    if "广告名称" in df.columns:
        df = df[df["广告名称"].astype(str).ne("总计")]
    if "消耗" in df.columns:
        df = df.sort_values("消耗", ascending=False)
    df = df.head(100)
    columns = [
        "REDACTED_DIM_REGION_TIER",
        "REDACTED_DIM_PLATFORM",
        "投放账户",
        "广告组ID",
        "广告ID",
        "广告名称",
        "REDACTED_CREATIVE_TYPE",
        "消耗",
        "REDACTED_CONVERSION_A_COUNT",
        "REDACTED_CONVERSION_B_COUNT",
        "REDACTED_CONVERSION_A_COST",
        "REDACTED_CONVERSION_B_COST",
        "REDACTED_CONVERSION_B_RATE",
        "REDACTED_CONVERSION_B_TO_C_RATE",
        "REDACTED_CONVERSION_C_TO_PAID_RATE",
        "当月ROI2",
        "滚动ROI2",
        "CPM",
        "CTR",
        "CVR",
    ]
    result = pick_columns(df, columns)
    result["备注占位"] = ""
    return add_context(result, "FBcreative_asset表现")


def candidate_type(row: pd.Series) -> str:
    waste = pd.to_numeric(row.get("空耗金额"), errors="coerce")
    spend = pd.to_numeric(row.get("消耗"), errors="coerce")
    effect = pd.to_numeric(row.get("成效"), errors="coerce")
    effect_cost = pd.to_numeric(row.get("成效成本"), errors="coerce")
    if pd.notna(waste) and waste > 0:
        return "空耗金额较高"
    if pd.notna(spend) and spend > 0 and (pd.isna(effect) or effect <= 0):
        return "高消耗低成效"
    if pd.notna(effect_cost) and effect_cost > 0:
        return "高成效成本"
    return "字段待校准"


def waste_sheet(material: pd.DataFrame) -> pd.DataFrame:
    df = material.copy()
    df.columns = [clean_column(c) for c in df.columns]
    df = df.loc[:, ~df.columns.duplicated()]
    for col in ["消耗", "成效", "成效成本", "空耗金额", "空耗金额占比"]:
        if col in df.columns:
            df[col] = to_number(df[col])
    mask = pd.Series(False, index=df.index)
    for col in ["空耗金额", "成效成本", "消耗"]:
        if col in df.columns:
            mask = mask | (df[col].fillna(0) > 0)
    df = df[mask].copy()
    if "广告名称" in df.columns:
        df = df[df["广告名称"].astype(str).ne("总计")]
    df["候选类型"] = df.apply(candidate_type, axis=1)
    sort_cols = [col for col in ["空耗金额", "成效成本", "消耗"] if col in df.columns]
    if sort_cols:
        df = df.sort_values(sort_cols, ascending=[False] * len(sort_cols))
    df = df.head(100)
    columns = ["REDACTED_DIM_REGION_TIER", "REDACTED_DIM_PLATFORM", "广告组ID", "广告ID", "广告名称", "REDACTED_CREATIVE_TYPE", "消耗", "成效", "成效成本", "空耗金额", "空耗金额占比", "候选类型"]
    result = pick_columns(df, columns)
    result.insert(0, "对象层级", "广告/creative_asset")
    result["人工确认位"] = ""
    return add_context(result, "FB空耗高成本候选")


def hkmo_sheet(hkmo: pd.DataFrame) -> pd.DataFrame:
    df = hkmo.copy()
    df.columns = [clean_column(c) for c in df.columns]
    df = df.loc[:, ~df.columns.duplicated()]
    keep_mask = pd.Series(True, index=df.index)
    if "creative_asset" in df.columns:
        keep_mask = df["creative_asset"].astype(str).str.contains("总计|图片|视频|creative_asset", na=False)
    df = df[keep_mask].copy()
    columns = [
        "REDACTED_DIM_OWNER_GROUP",
        "REDACTED_DIM_FORM_CHANNEL",
        "REDACTED_DIM_REGION_TIER",
        "REDACTED_DIM_TEST_ITEM",
        "REDACTED_FLAG_B",
        "REDACTED_FLAG_A",
        "creative_asset",
        "曝光",
        "点击",
        "点击费用",
        "CPM",
        "CTR",
        "CVR",
        "IPM",
        "消耗",
        "REDACTED_CONVERSION_A_COUNT",
        "REDACTED_CONVERSION_A_COST",
        "REDACTED_CONVERSION_B_COST",
        "正价课成本",
        "REDACTED_CONVERSION_B_COUNT",
        "当月REDACTED_CONVERSION_C_COUNT",
        "当月REDACTED_PAID_EVENT",
        "当月GMV",
        "当月ROI2",
    ]
    result = pick_columns(df, columns).rename(
        columns={"当月REDACTED_CONVERSION_C_COUNT": "REDACTED_CONVERSION_C_COUNT", "当月REDACTED_PAID_EVENT": "REDACTED_PAID_EVENT", "当月GMV": "GMV", "当月ROI2": "ROI2"}
    )
    result["备注占位"] = ""
    return add_context(result, "FBREDACTED_SEGMENT_TEST")


def gaps_sheet() -> pd.DataFrame:
    rows = [
        ("FB整体达成", "目标", "目标字段来源未接入", "无法自动计算达成率", "否", "投放/数据运营", "V0保留目标待接入"),
        ("FB整体达成", "达成率", "依赖目标字段", "无法自动填充", "否", "投放/数据运营", "不编造目标"),
        ("FBcreative_asset表现", "成效/成效成本", "口径待确认", "影响空耗和高成本候选解释", "否", "投放/数据运营", "V0只做候选标签"),
        ("FB渠道日监控", "消耗 vs 消耗不含CPT", "不同报表字段名称不一致", "影响成本口径", "否", "数据运营", "需确认周报采用哪个消耗口径"),
        ("FBcreative_asset表现", "当月ROI2 vs 滚动ROI2", "两个ROI2同时存在", "影响creative_asset表现排序", "否", "投放负责人", "V0两列都保留"),
        ("FB空耗高成本候选", "空耗阈值", "阈值未确认", "无法判断是否应处理", "否", "投放负责人", "V0按排序列候选，不给建议"),
        ("FBREDACTED_SEGMENT_TEST", "REDACTED_CREATIVE_TYPE/策略标签", "REDACTED_INTERNAL_TAG口径待确认", "影响区域/策略测试解读", "否", "投放/creative_asset负责人", "V0保留原字段"),
        ("投放师观点占位", "原因/动作", "V0不自动生成", "需要人工补充", "否", "投放师", "符合V0范围"),
    ]
    return pd.DataFrame(
        rows,
        columns=["周报模块", "缺口字段", "当前状态", "对周报影响", "是否阻塞V0", "需要谁确认", "备注"],
    )


def write_excel(sheets: dict[str, pd.DataFrame]) -> None:
    OUTPUT.parent.mkdir(parents=True, exist_ok=True)
    with pd.ExcelWriter(OUTPUT, engine="openpyxl") as writer:
        for name, df in sheets.items():
            df.to_excel(writer, sheet_name=name, index=False)
        for ws in writer.book.worksheets:
            ws.freeze_panes = "A2"
            ws.auto_filter.ref = ws.dimensions
            header_fill = PatternFill("solid", fgColor="1F4E78")
            header_font = Font(color="FFFFFF", bold=True)
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            for row in ws.iter_rows(min_row=2):
                for cell in row:
                    cell.alignment = Alignment(vertical="top", wrap_text=True)
            for column_cells in ws.columns:
                max_len = max(len(str(cell.value)) if cell.value is not None else 0 for cell in column_cells[:80])
                ws.column_dimensions[column_cells[0].column_letter].width = min(max(max_len + 2, 10), 34)


def write_markdown(sheets: dict[str, pd.DataFrame]) -> None:
    lines = [
        "# FB 自动周报 V0 验收表说明",
        "",
        "生成日期：2026-05-18",
        "",
        "## 当前产物",
        "",
        f"- Excel：`{OUTPUT.relative_to(ROOT)}`",
        "- 目标：直接给业务验收的汇报表，不再让用户阅读工程细底表。",
        "- 范围：只做 FB，不生成投放师观点，不自动写原因分析，不给调优建议。",
        "",
        "## Sheet 对应关系",
        "",
        "| Sheet | 对应飞书周报部分 | 行数 | 说明 |",
        "|---|---|---:|---|",
    ]
    notes = {
        "目录_对应飞书位置": "目录对照与校准提示。",
        "FB整体达成": "目标字段未接入，保留目标待接入。",
        "FB渠道日监控": "由渠道日监控长表整理成汇报表。",
        "FB链路类型日监控": "由链路类型日监控整理为日期 x 链路类型。",
        "FBcreative_asset表现": "按消耗排序保留Top 100有效creative_asset/广告记录。",
        "FB空耗高成本候选": "只列候选标签，不输出建议。",
        "FBREDACTED_SEGMENT_TEST": "保留区域/策略/REDACTED_CREATIVE_TAGS，供验收是否纳入V0。",
        "数据缺口与人工补充": "集中列出口径缺口和人工补充位。",
    }
    for name, df in sheets.items():
        lines.append(f"| `{name}` | {FEISHU[name]} | {len(df)} | {notes[name]} |")
    lines.extend(
        [
            "",
            "## 需要验收/校准",
            "",
            "1. `FB整体达成`：目标字段和达成率来源是否另有目标表。",
            "2. `FB渠道日监控`：分组名称是否就是周报正式分组，消耗是否采用含CPT或不含CPT。",
            "3. `FBcreative_asset表现`：REDACTED_CREATIVE_TYPE、广告名称、广告ID、广告组ID是否够用作周报creative_asset展示。",
            "4. `FB空耗高成本候选`：空耗阈值和成效/成效成本口径。",
            "5. `FBREDACTED_SEGMENT_TEST`：是否进入V0，还是留到V1策略分析。",
            "",
            "## 当前不能自动生成",
            "",
            "- 投放师观点、原因分析、下周动作。",
            "- 目标值和达成率，除非接入目标表。",
            "- 空耗是否应暂停/放量的最终判断。",
            "- REDACTED_CREATIVE_STRATEGY_TAGS的业务含义。",
        ]
    )
    REPORT.write_text("\n".join(lines) + "\n", encoding="utf-8")


def main() -> None:
    if not BASE.exists():
        raise SystemExit(f"Base workbook not found: {BASE}")
    channel = long_to_wide(read_base("channel_daily_long"))
    link_type = long_to_wide(read_base("link_type_daily_long"))
    material = read_base("material_wide")
    hkmo = read_base("hkmo_strategy_wide")

    sheets = {
        "目录_对应飞书位置": directory_sheet(),
        "FB整体达成": overall_sheet(channel),
        "FB渠道日监控": channel_sheet(channel),
        "FB链路类型日监控": link_type_sheet(link_type),
        "FBcreative_asset表现": material_sheet(material),
        "FB空耗高成本候选": waste_sheet(material),
        "FBREDACTED_SEGMENT_TEST": hkmo_sheet(hkmo),
        "数据缺口与人工补充": gaps_sheet(),
    }
    write_excel(sheets)
    write_markdown(sheets)
    print(OUTPUT)
    print(REPORT)


if __name__ == "__main__":
    main()
