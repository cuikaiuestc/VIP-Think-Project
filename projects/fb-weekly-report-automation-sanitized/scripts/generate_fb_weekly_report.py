#!/usr/bin/env python3
from __future__ import annotations

import argparse
import csv
import json
import math
from dataclasses import dataclass
from datetime import datetime
from html import escape
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


ROOT = Path(__file__).resolve().parents[1]
DEFAULT_REGISTRY = ROOT / "data/sample_registry.json"
DEFAULT_MANUAL = ROOT / "data/manual_material_supplement_template.csv"
DEFAULT_OUT_DIR = ROOT / "outputs/generated"

CORE_WEEKLY_SHEETS = {
    "02_RegionA_overall",
    "03_RegionA_link_type",
    "04_Material_type_MTD",
    "05_Recent_materials",
    "06_Waste_materials",
    "07_Other_regions",
    "08_Other_region_details",
}

MANUAL_FIELDS = [
    "section",
    "source_sheet_id",
    "source_row",
    "投放地区",
    "素材类型",
    "广告名称",
    "素材展示链接",
    "素材图片文件",
    "人工备注",
]


@dataclass
class ReportSection:
    sheet_name: str
    title: str
    rows: list[list[str]]
    source_type: str
    source_detail: str
    filters: str
    row_count: int
    status: str
    note: str = ""


def rel(path: Path) -> str:
    try:
        return str(path.relative_to(ROOT))
    except ValueError:
        return str(path)


def resolve_project_path(path_text: str | None) -> Path:
    if not path_text:
        return ROOT
    path = Path(path_text)
    return path if path.is_absolute() else ROOT / path


def load_json(path: Path) -> dict[str, Any]:
    if not path.exists():
        raise SystemExit(f"JSON file not found: {path}")
    return json.loads(path.read_text(encoding="utf-8"))


def read_csv_rows(path: Path) -> list[dict[str, str]]:
    if not path.exists():
        raise SystemExit(f"CSV file not found: {path}")
    with path.open(encoding="utf-8-sig", newline="") as handle:
        return list(csv.DictReader(handle))


def to_float(value: Any) -> float | None:
    if value is None:
        return None
    text = str(value).strip()
    if text in {"", "-", "nan", "None"}:
        return None
    if text.endswith("%"):
        text = text[:-1]
        try:
            return float(text) / 100
        except ValueError:
            return None
    try:
        number = float(text)
    except ValueError:
        return None
    if math.isnan(number):
        return None
    return number


def sum_float(values: list[Any]) -> float:
    return sum(value for value in (to_float(item) for item in values) if value is not None)


def fmt_number(value: float | None, decimals: int = 0) -> str:
    if value is None:
        return "-"
    return f"{value:,.{decimals}f}"


def fmt_rate(value: float | None, decimals: int = 2) -> str:
    if value is None:
        return "-"
    return f"{value * 100:.{decimals}f}%"


def fmt_metric(value: Any, metric: str) -> str:
    if isinstance(value, str):
        number = to_float(value)
        if number is None:
            return value
        value = number
    if value is None:
        return "-"
    if any(key in metric for key in ["率", "占比", "进度", "CTR", "CVR"]):
        return fmt_rate(float(value))
    if "ROI" in metric:
        return fmt_number(float(value), 2)
    if any(key in metric for key in ["成本", "消耗", "预算", "GMV", "金额", "GAP"]):
        return fmt_number(float(value), 2)
    if metric in {"例子数", "约课数", "记录数", "成效", "row_count"}:
        return fmt_number(float(value), 0)
    return str(value)


def long_metric(rows: list[dict[str, str]], period: str, group: str) -> dict[str, float | None]:
    out: dict[str, float | None] = {}
    for row in rows:
        if row.get("period") != period or row.get("business_date") != "total" or row.get("group") != group:
            continue
        metric = row.get("metric", "").strip()
        if metric:
            out[metric] = to_float(row.get("numeric_value") or row.get("value"))
    return out


def spend_from_metric_map(metrics: dict[str, float | None]) -> float | None:
    return metrics.get("消耗") or metrics.get("消耗 (不含CPT)")


def channel_summary(rows: list[dict[str, str]], period: str, groups: list[str], label: str) -> dict[str, Any]:
    group_metrics = [long_metric(rows, period, group) for group in groups]
    spend = sum_float([spend_from_metric_map(item) for item in group_metrics])
    leads = sum_float([item.get("例子数") for item in group_metrics])
    bookings = sum_float([item.get("约课数") for item in group_metrics])
    return {
        "周期": period,
        "分组": label,
        "消耗": spend,
        "例子数": leads,
        "例子成本": spend / leads if leads else None,
        "约课数": bookings,
        "约课成本": spend / bookings if bookings else None,
        "例子约课率": bookings / leads if leads else None,
    }


def single_group_summary(rows: list[dict[str, str]], period: str, group: str) -> dict[str, Any]:
    metrics = long_metric(rows, period, group)
    spend = spend_from_metric_map(metrics)
    leads = metrics.get("例子数")
    bookings = metrics.get("约课数")
    return {
        "周期": period,
        "分组": group,
        "消耗": spend,
        "例子数": leads,
        "例子成本": spend / leads if spend is not None and leads else None,
        "约课数": bookings,
        "约课成本": spend / bookings if spend is not None and bookings else None,
        "例子约课率": bookings / leads if leads else None,
        "CPM": metrics.get("CPM"),
        "CTR": metrics.get("CTR"),
        "CVR": metrics.get("CVR"),
    }


def dicts_to_rows(dict_rows: list[dict[str, Any]], headers: list[str]) -> list[list[str]]:
    output = [headers]
    for item in dict_rows:
        output.append([fmt_metric(item.get(header), header) for header in headers])
    return output


def material_base(rows: list[dict[str, str]], period: str = "month_to_date") -> list[dict[str, str]]:
    return [
        row
        for row in rows
        if row.get("period") == period
        and row.get("广告名称") not in {"", "总计"}
        and row.get("素材类型") not in {"", "总计"}
        and row.get("区域等级") not in {"", "总计"}
    ]


def material_aggregate(rows: list[dict[str, str]], group_fields: list[str]) -> list[dict[str, Any]]:
    grouped: dict[tuple[str, ...], list[dict[str, str]]] = {}
    for row in rows:
        grouped.setdefault(tuple(row.get(field, "") for field in group_fields), []).append(row)
    output: list[dict[str, Any]] = []
    for key, items in grouped.items():
        spend = sum_float([item.get("主投 | 消耗") for item in items])
        leads = sum_float([item.get("例子数") for item in items])
        bookings = sum_float([item.get("约课数") for item in items])
        monthly_gmv = sum_float([item.get("当月GMV") for item in items])
        rolling_gmv = sum_float([item.get("滚动GMV") for item in items])
        waste = sum_float([item.get("空耗金额") for item in items])
        row: dict[str, Any] = {field: value for field, value in zip(group_fields, key)}
        row.update(
            {
                "记录数": float(len(items)),
                "消耗": spend,
                "例子数": leads,
                "例子成本": spend / leads if leads else None,
                "约课数": bookings,
                "约课成本": spend / bookings if bookings else None,
                "约课率": bookings / leads if leads else None,
                "当月GMV": monthly_gmv or None,
                "当月ROI2": monthly_gmv / spend if spend and monthly_gmv else None,
                "滚动GMV": rolling_gmv or None,
                "滚动ROI2": rolling_gmv / spend if spend and rolling_gmv else None,
                "空耗金额": waste or None,
                "空耗金额占比": waste / spend if spend and waste else None,
            }
        )
        output.append(row)
    output.sort(key=lambda item: to_float(item.get("消耗")) or 0, reverse=True)
    return output


def recent_material_rows(rows: list[dict[str, str]], target_month: str) -> list[dict[str, Any]]:
    output: list[dict[str, Any]] = []
    for row in material_base(rows):
        spend = to_float(row.get("主投 | 消耗")) or 0
        leads = to_float(row.get("例子数")) or 0
        bookings = to_float(row.get("约课数")) or 0
        if row.get("区域等级") != "Region A" or row.get("creative_launch_month") != target_month:
            continue
        if spend < 1000 and leads < 5:
            continue
        output.append(
            {
                "投放地区": row.get("区域等级"),
                "投放平台": row.get("投放平台"),
                "素材类型": row.get("素材类型"),
                "广告名称": row.get("广告名称"),
                "清洗产出月份": row.get("creative_launch_month"),
                "消耗": spend,
                "例子数": leads,
                "例子成本": to_float(row.get("例子成本")) or (spend / leads if leads else None),
                "约课数": bookings,
                "约课成本": to_float(row.get("约课成本")) or (spend / bookings if bookings else None),
                "约课率": to_float(row.get("约课率")) or (bookings / leads if leads else None),
                "约课到课率": to_float(row.get("约课到课率")),
                "到课转化率": to_float(row.get("到课转化率")),
                "当月ROI2": to_float(row.get("当月ROI2")),
                "字段风险": "launch month is provided by sample data; production integration must validate the source field.",
            }
        )
    output.sort(key=lambda item: (item.get("消耗") or 0), reverse=True)
    return output


def waste_material_rows(rows: list[dict[str, str]]) -> list[dict[str, Any]]:
    output: list[dict[str, Any]] = []
    for row in material_base(rows):
        spend = to_float(row.get("主投 | 消耗")) or 0
        waste = to_float(row.get("空耗金额")) or 0
        if spend < 1000 or waste <= 0:
            continue
        output.append(
            {
                "投放地区": row.get("区域等级"),
                "投放平台": row.get("投放平台"),
                "素材类型": row.get("素材类型"),
                "广告名称": row.get("广告名称"),
                "消耗": spend,
                "空耗金额": waste,
                "空耗金额占比": to_float(row.get("空耗金额占比")) or (waste / spend if spend else None),
                "例子数": to_float(row.get("例子数")),
                "例子成本": to_float(row.get("例子成本")),
                "约课数": to_float(row.get("约课数")),
                "约课成本": to_float(row.get("约课成本")),
                "当月ROI2": to_float(row.get("当月ROI2")),
            }
        )
    output.sort(key=lambda item: item.get("空耗金额") or 0, reverse=True)
    return output


def table_from_fact_rows(rows: list[dict[str, str]], headers: list[str], limit: int | None = None) -> list[list[str]]:
    output = [headers]
    for row in rows[: limit or len(rows)]:
        output.append([fmt_metric(row.get(header), header) for header in headers])
    return output


def production_status(status: str) -> str:
    if status == "smartbi":
        return "auto"
    if status in {"smartbi_partial", "smartbi_cleaned"}:
        return "partial"
    if status in {"manual", "manual_contract"}:
        return "manual"
    if "blocked" in status:
        return "blocked"
    return "partial"


def production_readiness(sections: list[ReportSection]) -> dict[str, Any]:
    core = [section for section in sections if section.sheet_name in CORE_WEEKLY_SHEETS]
    counts = {"auto": 0, "partial": 0, "manual": 0, "blocked": 0}
    module_rows = []
    for section in core:
        status = production_status(section.status)
        counts[status] += 1
        module_rows.append(
            {
                "module": section.title,
                "sheet": section.sheet_name,
                "production_status": status,
                "source_type": section.source_type,
                "source_detail": section.source_detail,
                "filters": section.filters,
                "row_count": section.row_count,
                "known_gap": section.note,
            }
        )
    total = len(core) or 1
    readiness = "blocked" if counts["blocked"] else "ready_with_caveat" if counts["partial"] or counts["manual"] else "ready"
    return {
        "readiness": readiness,
        "core_module_count": len(core),
        "status_counts": counts,
        "auto_coverage": round(counts["auto"] / total, 4),
        "automated_or_partial_coverage": round((counts["auto"] + counts["partial"]) / total, 4),
        "modules": module_rows,
    }


def production_status_rows(readiness: dict[str, Any]) -> list[list[str]]:
    rows = [["模块", "输出sheet", "1.0状态", "来源", "输入行数", "缺口/人工边界"]]
    for item in readiness["modules"]:
        rows.append(
            [
                item["module"],
                item["sheet"],
                item["production_status"],
                item["source_type"],
                str(item["row_count"]),
                item["known_gap"],
            ]
        )
    counts = readiness["status_counts"]
    rows.append(
        [
            "汇总",
            "-",
            readiness["readiness"],
            f"auto={counts['auto']}; partial={counts['partial']}; manual={counts['manual']}; blocked={counts['blocked']}",
            str(readiness["core_module_count"]),
            f"auto_coverage={readiness['auto_coverage']:.0%}; auto_or_partial={readiness['automated_or_partial_coverage']:.0%}",
        ]
    )
    return rows


def module_contract_rows() -> list[list[str]]:
    return [
        ["模块", "分析问题", "指标", "维度", "时间窗", "BI构建方式", "人工/blocked字段"],
        ["整体达成", "MTD 目标与核心指标是否达成。", "消耗、例子数、约课数、成本、达成率、ROI", "区域/渠道组", "month_to_date / previous_week", "channel fact + target fact", "CPT remains manual"],
        ["链路表现", "H5、messaging、form 哪个贡献更高。", "消耗、例子数、约课数、约课率、成本", "link type", "previous_week", "link type fact", "deep adgroup join remains out of scope"],
        ["素材表现", "素材类型、近期跑量素材、空耗素材。", "消耗、例子数、约课数、ROI2、空耗", "区域/素材类型/广告名", "month_to_date", "material fact aggregation", "material images and links remain manual"],
        ["新老计划空耗", "新计划/老计划空耗是否异常。", "新老计划消耗、成效、空耗金额、空耗占比", "account type/week", "month_to_date", "new old plan waste fact", "ad-level actions not included"],
        ["赋能/高潜标签", "广告组标签用于展示，不输出自动决策。", "row_count、is_enable、is_high_potential", "adgroup name", "month_to_date", "derived tags from normalized fact", "no user-level data"],
    ]


def blocked_rows(registry: dict[str, Any]) -> list[list[str]]:
    return [
        ["模块", "状态", "缺口", "解决方案"],
        ["CPT成本", "manual", "sample fact does not include CPT cost.", "Provide a separate approved CPT supplement table before calculating included-CPT metrics."],
        ["素材链接/图片", "manual", "BI fields are unstable for preview assets.", "Use manual supplement CSV and keep it outside automated metric calculation."],
        ["SmartBI刷新", "blocked", "This package only consumes approved local files.", "Production integration must run SmartBI DATA CLI outside this repository and write normalized facts."],
        ["FB问答助手/诊断大脑", "out_of_scope", "Not included in this sanitized package.", "Contact the project owner for full integration."],
        ["公开传播", "blocked", "This package is for internal display and collaboration only.", "Confirm repository visibility and owner approval before any push."],
    ] + [["registry_known_gap", "caveat", str(item), "See docs/SECURITY_BOUNDARY.md"] for item in registry.get("known_gaps", [])]


def lineage_rows(registry: dict[str, Any], sections: list[ReportSection]) -> list[list[str]]:
    rows = [["模块", "输出sheet", "1.0状态", "来源类型", "来源明细", "时间窗/过滤条件", "输入行数", "状态", "人工字段边界"]]
    for section in sections:
        rows.append(
            [
                section.title,
                section.sheet_name,
                production_status(section.status),
                section.source_type,
                section.source_detail,
                section.filters,
                str(section.row_count),
                section.status,
                section.note or "无",
            ]
        )
    rows.append(
        [
            "registry",
            "00_Readme",
            "auto",
            "sanitized sample registry",
            registry.get("source_manifest", ""),
            f"data_as_of={registry.get('data_as_of', '')}; coverage={json.dumps(registry.get('coverage', {}), ensure_ascii=False)}",
            "",
            registry.get("status", ""),
            "No SmartBI refresh or external write happens in this package.",
        ]
    )
    return rows


def build_sections(registry: dict[str, Any], manual_rows: list[dict[str, str]]) -> list[ReportSection]:
    facts = registry.get("facts", {})
    channel_path = resolve_project_path(facts.get("fb_channel_daily_monitor"))
    link_path = resolve_project_path(facts.get("fb_link_type_daily_monitor"))
    material_path = resolve_project_path(facts.get("fb_material_chain_metrics"))
    target_path = resolve_project_path(facts.get("fb_target_achievement"))
    new_old_path = resolve_project_path(facts.get("fb_new_old_plan_waste"))
    tag_path = resolve_project_path(facts.get("fb_adgroup_tags"))

    channel = read_csv_rows(channel_path)
    link = read_csv_rows(link_path)
    material = read_csv_rows(material_path)
    target = read_csv_rows(target_path)
    new_old = read_csv_rows(new_old_path)
    tags = read_csv_rows(tag_path)
    material_mtd = material_base(material)
    target_month = str(registry.get("data_as_of", ""))[:7] or "2026-05"

    sections: list[ReportSection] = [
        ReportSection(
            "01_Module_contract",
            "飞书分析逻辑到BI构建映射",
            module_contract_rows(),
            "contract",
            "Sanitized module contract derived from FB weekly 1.0 boundaries.",
            "Schema and module contract only.",
            0,
            "manual_contract",
            "用于约束后续 BI 构建，不复制真实指标值。",
        )
    ]

    hk_rows = [
        single_group_summary(channel, "month_to_date", "FB_REGION_A_KOL"),
        single_group_summary(channel, "month_to_date", "FB_REGION_A_STANDARD"),
        channel_summary(channel, "month_to_date", ["FB_REGION_A_KOL", "FB_REGION_A_STANDARD"], "FB_REGION_A_TOTAL"),
        single_group_summary(channel, "previous_week", "FB_REGION_A_KOL"),
        single_group_summary(channel, "previous_week", "FB_REGION_A_STANDARD"),
        channel_summary(channel, "previous_week", ["FB_REGION_A_KOL", "FB_REGION_A_STANDARD"], "FB_REGION_A_TOTAL"),
    ]
    sections.append(
        ReportSection(
            "02_RegionA_overall",
            "Region A 整体表现-BI",
            dicts_to_rows(hk_rows, ["周期", "分组", "消耗", "例子数", "例子成本", "约课数", "约课成本", "例子约课率", "CPM", "CTR", "CVR"]),
            "normalized_facts",
            f"sample channel fact -> {rel(channel_path)}",
            "period in [month_to_date, previous_week]; business_date=total",
            len(channel),
            "smartbi",
            "目标达成另见 target fact；CPT仍人工补充。",
        )
    )

    link_rows = [
        single_group_summary(link, "previous_week", "REGION_A_H5"),
        single_group_summary(link, "previous_week", "REGION_A_MESSAGING"),
        single_group_summary(link, "previous_week", "REGION_A_FORM"),
    ]
    sections.append(
        ReportSection(
            "03_RegionA_link_type",
            "Region A 链路表现-BI",
            dicts_to_rows(link_rows, ["周期", "分组", "消耗", "例子数", "例子成本", "约课数", "约课成本", "例子约课率", "CPM", "CTR", "CVR"]),
            "normalized_facts",
            f"sample link type fact -> {rel(link_path)}",
            "period=previous_week; group in [H5, messaging, form]",
            len(link),
            "smartbi_partial",
            "链路 x 广告组标签深度 join 留到完整版本。",
        )
    )

    sections.append(
        ReportSection(
            "04_Material_type_MTD",
            "素材类型表现-BI MTD",
            dicts_to_rows(
                material_aggregate(material_mtd, ["区域等级", "投放平台", "素材类型"]),
                ["区域等级", "投放平台", "素材类型", "记录数", "消耗", "例子数", "例子成本", "约课数", "约课成本", "约课率", "当月ROI2", "滚动ROI2", "空耗金额", "空耗金额占比"],
            ),
            "normalized_facts",
            f"sample material fact -> {rel(material_path)}",
            "period=month_to_date; group by region/platform/material_type",
            len(material_mtd),
            "smartbi",
            "素材链接/图片不在自动字段内。",
        )
    )

    sections.append(
        ReportSection(
            "05_Recent_materials",
            "近期跑量素材-BI",
            dicts_to_rows(
                recent_material_rows(material, target_month),
                ["投放地区", "投放平台", "素材类型", "广告名称", "清洗产出月份", "消耗", "例子数", "例子成本", "约课数", "约课成本", "约课率", "约课到课率", "到课转化率", "当月ROI2", "字段风险"],
            ),
            "normalized_facts + cleaned",
            f"sample material fact -> {rel(material_path)}",
            f"Region A; launch_month={target_month}; spend>=1000 or leads>=5",
            len(material_mtd),
            "smartbi_cleaned",
            "素材展示链接/图片仍由人工补充。",
        )
    )

    sections.append(
        ReportSection(
            "06_Waste_materials",
            "空耗素材-BI",
            dicts_to_rows(
                waste_material_rows(material),
                ["投放地区", "投放平台", "素材类型", "广告名称", "消耗", "空耗金额", "空耗金额占比", "例子数", "例子成本", "约课数", "约课成本", "当月ROI2"],
            ),
            "normalized_facts",
            f"sample material fact -> {rel(material_path)}",
            "period=month_to_date; spend>=1000; waste>0",
            len(material_mtd),
            "smartbi",
            "新老计划/广告组维度空耗见专门模块；素材图链仍人工。",
        )
    )

    region_rows = []
    for group in ["FB_REGION_B_STANDARD", "FB_REGION_C_STANDARD"]:
        region_rows.append(single_group_summary(channel, "month_to_date", group))
        region_rows.append(single_group_summary(channel, "previous_week", group))
    sections.append(
        ReportSection(
            "07_Other_regions",
            "其他区域表现-BI",
            dicts_to_rows(region_rows, ["周期", "分组", "消耗", "例子数", "例子成本", "约课数", "约课成本", "例子约课率", "CPM", "CTR", "CVR"]),
            "normalized_facts",
            f"sample channel fact -> {rel(channel_path)}",
            "period in [month_to_date, previous_week]; group in [Region B, Region C]",
            len(channel),
            "smartbi_partial",
            "区域映射在生产接入时需要 owner 确认。",
        )
    )

    detail_rows = []
    for row in material_mtd:
        if row.get("区域等级") not in {"Region B", "Region C"}:
            continue
        spend = to_float(row.get("主投 | 消耗")) or 0
        leads = to_float(row.get("例子数")) or 0
        if spend < 1000 and leads < 3:
            continue
        detail_rows.append(
            {
                "区域等级": row.get("区域等级"),
                "投放平台": row.get("投放平台"),
                "素材类型": row.get("素材类型"),
                "广告名称": row.get("广告名称"),
                "消耗": spend,
                "例子数": leads,
                "例子成本": to_float(row.get("例子成本")),
                "约课数": to_float(row.get("约课数")),
                "约课成本": to_float(row.get("约课成本")),
                "约课率": to_float(row.get("约课率")),
                "当月ROI2": to_float(row.get("当月ROI2")),
            }
        )
    sections.append(
        ReportSection(
            "08_Other_region_details",
            "其他区域素材明细-BI",
            dicts_to_rows(detail_rows, ["区域等级", "投放平台", "素材类型", "广告名称", "消耗", "例子数", "例子成本", "约课数", "约课成本", "约课率", "当月ROI2"]),
            "normalized_facts",
            f"sample material fact -> {rel(material_path)}",
            "Region B/C; spend>=1000 or leads>=3",
            len(material_mtd),
            "smartbi",
            "素材预览与CPT不在裁剪包内。",
        )
    )

    sections.append(
        ReportSection(
            "09_Target_achievement",
            "MTD目标达成-BI",
            table_from_fact_rows(
                target,
                ["主投", "辅投", "平台", "区域等级", "全月预算", "消耗MTD", "MTD消耗进度", "主投例子月目标", "主投例子数", "主投例子MTD达成率", "例子成本目标", "例子成本", "约课成本目标", "约课成本", "当月GMV目标", "当月GMV达成", "当月GMV达成率", "主投当月ROI目标", "主投当月ROI达成"],
            ),
            "normalized_facts",
            f"sample target fact -> {rel(target_path)}",
            "period=month_to_date; platform=FB",
            len(target),
            "smartbi",
            "用于目标值、达成率、ROI展示；不输出经营动作建议。",
        )
    )

    sections.append(
        ReportSection(
            "10_New_old_plan_waste",
            "新老计划空耗-BI",
            table_from_fact_rows(
                new_old,
                ["账户类型", "周次", "汇总 | 消耗", "汇总 | 成效", "汇总 | 成效成本", "新计划 | 消耗", "新计划 | 成效", "新计划 | 成效成本", "老计划 | 消耗", "老计划 | 成效", "老计划 | 成效成本", "广告组维度 | 空耗金额", "广告组维度 | 空耗金额占比", "广告维度 | 空耗金额", "广告维度 | 空耗金额占比"],
            ),
            "normalized_facts",
            f"sample new old plan waste fact -> {rel(new_old_path)}",
            "month_to_date; summary/latest week",
            len(new_old),
            "smartbi",
            "只做展示，不自动给出停投或加预算动作。",
        )
    )

    sections.append(
        ReportSection(
            "11_Adgroup_tags",
            "赋能/高潜广告组标签-BI",
            table_from_fact_rows(
                tags,
                ["区域等级", "三大区域", "渠道一级分类", "渠道二级分类", "末次渠道名称", "投放账户", "广告组名称", "row_count", "is_enable", "is_high_potential"],
            ),
            "normalized_facts + derived rule",
            f"sample adgroup tags fact -> {rel(tag_path)}",
            "aggregated adgroup-level rows only",
            len(tags),
            "smartbi",
            "不输出用户级明细。",
        )
    )

    sections.append(
        ReportSection(
            "12_Blocked_and_manual",
            "Blocked项与接入建议",
            blocked_rows(registry),
            "blocked/manual",
            "sample registry known_gaps + package boundary",
            "manual / partial / blocked boundary list",
            0,
            "manual_contract",
            "强脱敏裁剪包保留能力边界，不包装成完整生产系统。",
        )
    )

    sections.append(
        ReportSection(
            "13_Manual_materials",
            "人工素材补充位",
            table_from_fact_rows(manual_rows, MANUAL_FIELDS),
            "manual",
            rel(DEFAULT_MANUAL),
            "manual CSV for preview link/image/comment",
            len(manual_rows),
            "manual",
            "素材网址、图片文件、预览图由人工补充，不参与指标计算。",
        )
    )

    return sections


def md_escape(text: str) -> str:
    return str(text).replace("|", "\\|")


def markdown_table(rows: list[list[str]]) -> str:
    if not rows:
        return "_无数据_"
    width = max(len(row) for row in rows)
    normalized = [row + [""] * (width - len(row)) for row in rows]
    lines = [
        "| " + " | ".join(md_escape(cell) for cell in normalized[0]) + " |",
        "| " + " | ".join(["---"] * width) + " |",
    ]
    for row in normalized[1:]:
        lines.append("| " + " | ".join(md_escape(cell) for cell in row) + " |")
    return "\n".join(lines)


def write_markdown_report(registry: dict[str, Any], sections: list[ReportSection], registry_path: Path, out_path: Path) -> None:
    readiness = production_readiness(sections)
    parts = [
        "# FB周报自动化 1.0 脱敏示例报告",
        "",
        f"- generated_at: {datetime.now().isoformat(timespec='seconds')}",
        f"- data_as_of: {registry.get('data_as_of', '')}",
        f"- registry: `{rel(registry_path)}`",
        "- execution_boundary: read local sanitized sample facts only; no SmartBI refresh; no Feishu write-back; no Meta API call.",
        f"- readiness: `{readiness['readiness']}`",
        "",
        "## 0. 模块状态",
        "",
        markdown_table(production_status_rows(readiness)),
        "",
        "## 0.1 数据提取链路审计",
        "",
        markdown_table(lineage_rows(registry, sections)),
        "",
    ]
    for section in sections:
        parts.extend(
            [
                f"## {section.title}",
                "",
                f"- source_type: `{section.source_type}`",
                f"- source_detail: `{section.source_detail}`",
                f"- filters: {section.filters}",
                f"- status: `{section.status}`",
                f"- note: {section.note}",
                "",
                markdown_table(section.rows),
                "",
            ]
        )
    out_path.write_text("\n".join(parts), encoding="utf-8")


def safe_sheet_name(name: str, used: set[str]) -> str:
    invalid = set('[]:*?/\\')
    cleaned = "".join("_" if char in invalid else char for char in name).strip() or "Sheet"
    cleaned = cleaned[:31]
    candidate = cleaned
    suffix = 1
    while candidate in used:
        marker = f"_{suffix}"
        candidate = f"{cleaned[:31 - len(marker)]}{marker}"
        suffix += 1
    used.add(candidate)
    return candidate


def set_sheet_widths(ws, max_width: int = 48) -> None:
    for col_idx, column in enumerate(ws.columns, start=1):
        width = 10
        for cell in column:
            value = "" if cell.value is None else str(cell.value)
            width = max(width, min(max_width, len(value) + 2))
        ws.column_dimensions[get_column_letter(col_idx)].width = width


def style_table(ws, header_row: int, max_row: int, max_col: int) -> None:
    header_fill = PatternFill("solid", fgColor="D9EAF7")
    note_fill = PatternFill("solid", fgColor="FFF2CC")
    border = Border(
        left=Side(style="thin", color="D9D9D9"),
        right=Side(style="thin", color="D9D9D9"),
        top=Side(style="thin", color="D9D9D9"),
        bottom=Side(style="thin", color="D9D9D9"),
    )
    for row in ws.iter_rows(min_row=1, max_row=max_row, max_col=max_col):
        for cell in row:
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            cell.border = border
    for cell in ws[header_row]:
        cell.font = Font(bold=True)
        cell.fill = header_fill
    if max_row >= 2:
        for cell in ws[2]:
            cell.fill = note_fill
    ws.freeze_panes = f"A{header_row + 1}"


def write_excel_report(registry: dict[str, Any], sections: list[ReportSection], registry_path: Path, out_path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "00_Readme"
    intro_rows = [
        ["FB周报自动化 1.0 脱敏示例报告"],
        ["generated_at", datetime.now().isoformat(timespec="seconds")],
        ["data_as_of", registry.get("data_as_of", "")],
        ["registry", rel(registry_path)],
        ["execution_boundary", "local sanitized sample facts only; no external writes"],
    ]
    for row in intro_rows:
        ws.append(row)
    ws["A1"].font = Font(bold=True, size=14)
    set_sheet_widths(ws)

    used = {ws.title}
    readiness = production_readiness(sections)
    status_sheet = wb.create_sheet(safe_sheet_name("00_Module_status", used))
    status_sheet.append(["FB weekly report 1.0 module status"])
    status_sheet.append([f"readiness: {readiness['readiness']}", "Sanitized internal collaboration package."])
    status_sheet.append([])
    status_rows = production_status_rows(readiness)
    for row in status_rows:
        status_sheet.append(row)
    style_table(status_sheet, 4, status_sheet.max_row, max(len(row) for row in status_rows))
    set_sheet_widths(status_sheet, max_width=64)

    lineage_sheet = wb.create_sheet(safe_sheet_name("00_Lineage", used))
    lineage_sheet.append(["Data lineage audit"])
    lineage_sheet.append(["Reads only registry-declared local sample normalized facts."])
    lineage_sheet.append([])
    lineage_data = lineage_rows(registry, sections)
    for row in lineage_data:
        lineage_sheet.append(row)
    style_table(lineage_sheet, 4, lineage_sheet.max_row, max(len(row) for row in lineage_data))
    set_sheet_widths(lineage_sheet, max_width=64)

    for section in sections:
        sheet = wb.create_sheet(safe_sheet_name(section.sheet_name, used))
        sheet.append([section.title])
        sheet.append([section.source_detail])
        sheet.append([f"filters: {section.filters}", f"status: {section.status}", f"note: {section.note}"])
        rows = section.rows or [["无数据"]]
        for row in rows:
            sheet.append(row)
        max_col = max((len(row) for row in rows), default=1)
        style_table(sheet, 4, sheet.max_row, max_col)
        sheet["A1"].font = Font(bold=True, size=13)
        set_sheet_widths(sheet)
    wb.save(out_path)


def html_table(rows: list[list[str]]) -> str:
    if not rows:
        return "<p class=\"empty\">No data</p>"
    header = rows[0]
    body = rows[1:]
    thead = "".join(f"<th>{escape(str(cell))}</th>" for cell in header)
    body_rows = []
    for row in body:
        cells = "".join(f"<td>{escape(str(cell))}</td>" for cell in row)
        body_rows.append(f"<tr>{cells}</tr>")
    return f"<div class=\"table-wrap\"><table><thead><tr>{thead}</tr></thead><tbody>{''.join(body_rows)}</tbody></table></div>"


def write_html_report(registry: dict[str, Any], sections: list[ReportSection], registry_path: Path, out_path: Path) -> None:
    readiness = production_readiness(sections)
    display_sections = [
        ReportSection("00_Module_status", "模块状态", production_status_rows(readiness), "manifest", "computed module readiness", "core modules", len(sections), readiness["readiness"]),
        ReportSection("00_Lineage", "数据提取链路审计", lineage_rows(registry, sections), "lineage", "registry + section metadata", "all sections", len(sections), "audit"),
    ] + sections
    nav = "\n".join(f"<a href=\"#{escape(section.sheet_name)}\">{escape(section.title)}</a>" for section in display_sections)
    body = []
    for section in display_sections:
        body.append(
            f"""
            <section id="{escape(section.sheet_name)}">
              <h2>{escape(section.title)}</h2>
              <p class="source">source: {escape(section.source_detail)}</p>
              <p class="note">status: {escape(section.status)} | filters: {escape(section.filters)} | {escape(section.note)}</p>
              {html_table(section.rows)}
            </section>
            """
        )
    html = f"""<!doctype html>
<html lang="zh-CN">
<head>
  <meta charset="utf-8">
  <meta name="viewport" content="width=device-width, initial-scale=1">
  <title>FB周报自动化 1.0 脱敏示例报告</title>
  <style>
    :root {{
      color-scheme: light;
      --text: #1f2933;
      --muted: #607080;
      --line: #d9e2ec;
      --head: #d9eaf7;
      --note: #fff7d6;
      --bg: #f6f8fb;
      --panel: #ffffff;
      --accent: #2368a2;
    }}
    * {{ box-sizing: border-box; }}
    body {{ margin: 0; background: var(--bg); color: var(--text); font: 14px/1.55 -apple-system, BlinkMacSystemFont, "Segoe UI", sans-serif; }}
    header {{ padding: 28px 32px 18px; background: var(--panel); border-bottom: 1px solid var(--line); }}
    h1 {{ margin: 0 0 12px; font-size: 24px; letter-spacing: 0; }}
    .meta {{ display: grid; grid-template-columns: max-content 1fr; gap: 6px 14px; color: var(--muted); }}
    nav {{ display: flex; gap: 8px; overflow-x: auto; padding: 12px 32px; background: #edf4fa; border-bottom: 1px solid var(--line); position: sticky; top: 0; z-index: 1; }}
    nav a {{ flex: 0 0 auto; color: var(--accent); text-decoration: none; padding: 4px 8px; border: 1px solid var(--line); background: var(--panel); border-radius: 4px; font-size: 13px; }}
    main {{ padding: 20px 32px 40px; }}
    section {{ background: var(--panel); border: 1px solid var(--line); border-radius: 6px; margin: 0 0 18px; padding: 18px; }}
    h2 {{ margin: 0 0 6px; font-size: 18px; letter-spacing: 0; }}
    .source {{ margin: 0 0 8px; color: var(--muted); font-size: 12px; }}
    .note {{ margin: 0 0 12px; padding: 8px 10px; background: var(--note); border: 1px solid #f2dd9d; border-radius: 4px; }}
    .table-wrap {{ width: 100%; overflow-x: auto; border: 1px solid var(--line); }}
    table {{ width: max-content; min-width: 100%; border-collapse: collapse; background: #fff; }}
    th, td {{ border: 1px solid var(--line); padding: 6px 8px; text-align: left; vertical-align: middle; white-space: nowrap; }}
    th {{ background: var(--head); font-weight: 700; }}
    td {{ max-width: 420px; overflow: hidden; text-overflow: ellipsis; }}
    .empty {{ color: var(--muted); }}
    @media (max-width: 720px) {{
      header, main {{ padding-left: 16px; padding-right: 16px; }}
      nav {{ padding-left: 16px; padding-right: 16px; }}
      section {{ padding: 12px; }}
    }}
  </style>
</head>
<body>
  <header>
    <h1>FB周报自动化 1.0 脱敏示例报告</h1>
    <div class="meta">
      <strong>generated_at</strong><span>{escape(datetime.now().isoformat(timespec="seconds"))}</span>
      <strong>data_as_of</strong><span>{escape(str(registry.get("data_as_of", "")))}</span>
      <strong>registry</strong><span>{escape(rel(registry_path))}</span>
      <strong>readiness</strong><span>{escape(readiness["readiness"])}</span>
      <strong>boundary</strong><span>local sanitized sample facts only; no external writes</span>
    </div>
  </header>
  <nav>{nav}</nav>
  <main>{''.join(body)}</main>
</body>
</html>
"""
    out_path.write_text(html, encoding="utf-8")


def write_manifest(registry: dict[str, Any], sections: list[ReportSection], registry_path: Path, outputs: dict[str, Path], out_path: Path) -> dict[str, Any]:
    readiness = production_readiness(sections)
    manifest = {
        "schema_version": "fb-weekly-sanitized-production-manifest-v1",
        "generated_at": datetime.now().isoformat(timespec="seconds"),
        "target_version": "fb_weekly_report_1_0_sanitized",
        "readiness": readiness["readiness"],
        "data_as_of": registry.get("data_as_of", ""),
        "registry": rel(registry_path),
        "outputs": {key: rel(value) for key, value in outputs.items()},
        "module_status": readiness,
        "execution_boundary": {
            "smartbi_called": False,
            "feishu_written": False,
            "meta_api_called": False,
            "external_writes": False,
            "credential_read": False,
            "uses_mock_data": True,
        },
        "not_included": [
            "SmartBI DATA CLI implementation",
            "FB ad QA assistant or diagnosis brain",
            "real business data",
            "credentials and private config",
            "automatic ad account decisions",
        ],
    }
    out_path.write_text(json.dumps(manifest, ensure_ascii=False, indent=2), encoding="utf-8")
    return manifest


def main() -> None:
    parser = argparse.ArgumentParser(description="Generate sanitized FB weekly report 1.0 samples.")
    parser.add_argument("--registry", default=str(DEFAULT_REGISTRY), help="Path to sample registry JSON.")
    parser.add_argument("--manual", default=str(DEFAULT_MANUAL), help="Path to manual material supplement CSV.")
    parser.add_argument("--out-dir", default=str(DEFAULT_OUT_DIR), help="Output directory.")
    args = parser.parse_args()

    registry_path = resolve_project_path(args.registry)
    manual_path = resolve_project_path(args.manual)
    out_dir = resolve_project_path(args.out_dir)
    out_dir.mkdir(parents=True, exist_ok=True)

    registry = load_json(registry_path)
    manual_rows = read_csv_rows(manual_path)
    sections = build_sections(registry, manual_rows)

    report_date = str(registry.get("data_as_of") or datetime.now().date().isoformat())
    markdown_path = out_dir / f"FB_weekly_report_sanitized_{report_date}.md"
    excel_path = out_dir / f"FB_weekly_report_sanitized_{report_date}.xlsx"
    html_path = out_dir / f"FB_weekly_report_sanitized_{report_date}.html"
    manifest_path = out_dir / f"FB_weekly_report_sanitized_{report_date}_manifest.json"

    write_markdown_report(registry, sections, registry_path, markdown_path)
    write_excel_report(registry, sections, registry_path, excel_path)
    write_html_report(registry, sections, registry_path, html_path)
    manifest = write_manifest(
        registry,
        sections,
        registry_path,
        {"markdown": markdown_path, "excel": excel_path, "html": html_path, "manifest": manifest_path},
        manifest_path,
    )
    print(json.dumps({"outputs": manifest["outputs"], "readiness": manifest["readiness"]}, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()

